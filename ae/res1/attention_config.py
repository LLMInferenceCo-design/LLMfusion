import sys
import os
import copy
import pandas as pd
# 获取当前工作目录
if __name__ == "__main__":
    current_dir = os.getcwd()
    print("当前目录:", current_dir)

    # 更改到上两级目录
    parent_dir = os.path.dirname(os.path.dirname(current_dir))
    os.chdir(parent_dir)
    sys.path.append(parent_dir)

    # 验证当前工作目录
    new_dir = os.getcwd()
    print("更改后的目录:", new_dir)
# sys.path.append("/root/paper/LLMfusion")
# os.chdir("/root/paper/LLMfusion")
import logging
from ae.fig1.change_size import change_hardware_params
from software_model.matmul_horizontal_fusion import HorizontalMatmulFusion
from software_model.mutmul_fusion import MatmulFusion
from software_model.matmul import Matmul, BatchedMatmul
from software_model.DataFrame import DataType, Tensor, data_type_dict
from hardware_model.system import System
from software_model.softmax import Softmax
from software_model.flash_attention_fusion import FlashAttentionFusion
from LLM_model.opt175b import opt175b_prefill
import json
import time
from loguru import logger
from util.mapping import Mapping
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook

config = ['A', 'B', 'C', 'D', 'E']
core_count = [8, 32, 128, 256, 512]
sublane_count = 4
vector_width = [512, 128, 32, 16, 8]
array_height = [64, 32, 16, 8, 4]
array_width = [64, 32, 16, 16, 16]
SRAM_KB = [3072, 768, 192, 96, 48]

def change_config_prefill_test(config, batch_size:int, Lin:int, device_count:int):
    Nhead = 96
    # batch_size = 1
    M = Lin
    N = 128
    kv_cache = 0

    assert Nhead % device_count == 0
    batch = batch_size * Nhead//device_count
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(config, arch_specs)

    QK = BatchedMatmul(data_type=data_type_dict['fp16'])
    _ = QK(Tensor([batch, M, N]), Tensor([batch, N, M + kv_cache]))

    S = Softmax(data_type=data_type_dict['fp16'])
    _ = S(Tensor([batch * M, M + kv_cache]))

    SV = BatchedMatmul(data_type=data_type_dict['fp16'])
    _ = SV(Tensor([batch, M, M + kv_cache]), Tensor([batch, M + kv_cache, N]))

    flash_attention = FlashAttentionFusion([QK, S, SV], data_type_dict['fp16'])
    time_s = flash_attention.compile_and_simulate(system.device)
    # tmp = flash_attention.simulate(flash_attention.best_mapping, system.device)
    clock = time_s * system.device.compute_module.clock_freq
    # logger.debug(f"Prefill config: {config['config']}, batch_size: {batch_size}, Lin: {Lin}, device_count: {device_count}, clock num: {clock}")
    return time_s, area, clock

def change_config_decode_test(config, batch_size:int, Lin:int, device_count:int):
    Nhead = 96
    # batch_size = 1
    M = 1
    N = 128
    kv_cache = Lin
    assert Nhead % device_count == 0
    batch = batch_size * Nhead//device_count
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(config, arch_specs)

    QK = BatchedMatmul(data_type=data_type_dict['fp16'])
    _ = QK(Tensor([batch, M, N]), Tensor([batch, N, M + kv_cache]))

    S = Softmax(data_type=data_type_dict['fp16'])
    _ = S(Tensor([batch * M, M + kv_cache]))

    SV = BatchedMatmul(data_type=data_type_dict['fp16'])
    _ = SV(Tensor([batch, M, M + kv_cache]), Tensor([batch, M + kv_cache, N]))

    flash_attention = FlashAttentionFusion([QK, S, SV], data_type_dict['fp16'])
    time_s = flash_attention.compile_and_simulate(system.device)
    # tmp = flash_attention.simulate(flash_attention.best_mapping, system.device)
    clock = time_s * system.device.compute_module.clock_freq
    # logger.debug(f"Prefill config: {config['config']}, batch_size: {batch_size}, Lin: {Lin}, device_count: {device_count}, clock num: {clock}")
    return time_s, area, clock


def change_config_prefill_no_V_test(config, batch_size: int, Lin: int, device_count: int):
    Nhead = 96
    # batch_size = 1
    M = Lin
    N = 128
    kv_cache = 0

    assert Nhead % device_count == 0
    batch = batch_size * Nhead // device_count
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(config, arch_specs)

    mul1 = Matmul(data_type=data_type_dict['fp16'])

    _ = mul1(Tensor([M, N]), Tensor([N, M + kv_cache]))

    mul1_fusion = MatmulFusion([mul1], data_type_dict['fp16'])

    mul_fusion = [copy.deepcopy(mul1_fusion) for _ in range(batch)]

    mul1 = HorizontalMatmulFusion(mul_fusion, data_type_dict['fp16'])
    QK_times = mul1.compile_and_simulate(system.device)
    QK_clock = QK_times * system.device.compute_module.clock_freq

    S = Softmax(data_type=data_type_dict['fp16'])
    _ = S(Tensor([batch * M, M + kv_cache]))
    S_times = S.compile_and_simulate(system.device)
    S_clock = S_times * system.device.compute_module.clock_freq

    mul2 = Matmul(data_type=data_type_dict['fp16'])
    _ = mul2(Tensor([M, M + kv_cache]), Tensor([M + kv_cache, N]))
    mul2_fusion = MatmulFusion([mul2], data_type_dict['fp16'])
    mul_fusion = [copy.deepcopy(mul2_fusion) for _ in range(batch)]
    mul2 = HorizontalMatmulFusion(mul_fusion, data_type_dict['fp16'])
    SV_times = mul2.compile_and_simulate(system.device)
    SV_clock = SV_times * system.device.compute_module.clock_freq

    time_s = QK_times + S_times + SV_times
    clock = QK_clock + S_clock + SV_clock

    return time_s, area, clock


def change_config_decode_no_V_test(config, batch_size: int, Lin: int, device_count: int):
    Nhead = 96
    # batch_size = 1
    M = 1
    N = 128
    kv_cache = Lin

    assert Nhead % device_count == 0
    batch = batch_size * Nhead // device_count
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(config, arch_specs)

    mul1 = Matmul(data_type=data_type_dict['fp16'])

    _ = mul1(Tensor([M, N]), Tensor([N, M + kv_cache]))

    mul1_fusion = MatmulFusion([mul1], data_type_dict['fp16'])

    mul_fusion = [copy.deepcopy(mul1_fusion) for _ in range(batch)]

    mul1 = HorizontalMatmulFusion(mul_fusion, data_type_dict['fp16'])
    QK_times = mul1.compile_and_simulate(system.device)
    QK_clock = QK_times * system.device.compute_module.clock_freq

    S = Softmax(data_type=data_type_dict['fp16'])
    _ = S(Tensor([batch * M, M + kv_cache]))
    S_times = S.compile_and_simulate(system.device)
    S_clock = S_times * system.device.compute_module.clock_freq

    mul2 = Matmul(data_type=data_type_dict['fp16'])
    _ = mul2(Tensor([M, M + kv_cache]), Tensor([M + kv_cache, N]))
    mul2_fusion = MatmulFusion([mul2], data_type_dict['fp16'])
    mul_fusion = [copy.deepcopy(mul2_fusion) for _ in range(batch)]
    mul2 = HorizontalMatmulFusion(mul_fusion, data_type_dict['fp16'])
    SV_times = mul2.compile_and_simulate(system.device)
    SV_clock = SV_times * system.device.compute_module.clock_freq

    time_s = QK_times + S_times + SV_times
    clock = QK_clock + S_clock + SV_clock

    return time_s, area, clock

def change_config_prefill_no_V_test(config, batch_size: int, Lin: int, device_count: int):
    Nhead = 96
    # batch_size = 1
    M = Lin
    N = 128
    kv_cache = 0

    assert Nhead % device_count == 0
    batch = batch_size * Nhead // device_count
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(config, arch_specs)

    mul1 = Matmul(data_type=data_type_dict['fp16'])

    _ = mul1(Tensor([M, N]), Tensor([N, M + kv_cache]))

    mul1_fusion = MatmulFusion([mul1], data_type_dict['fp16'])

    mul_fusion = [copy.deepcopy(mul1_fusion) for _ in range(batch)]

    mul1 = HorizontalMatmulFusion(mul_fusion, data_type_dict['fp16'])
    QK_times = mul1.compile_and_simulate(system.device)
    QK_clock = QK_times * system.device.compute_module.clock_freq

    S = Softmax(data_type=data_type_dict['fp16'])
    _ = S(Tensor([batch * M, M + kv_cache]))
    S_times = S.compile_and_simulate(system.device)
    S_clock = S_times * system.device.compute_module.clock_freq

    mul2 = Matmul(data_type=data_type_dict['fp16'])
    _ = mul2(Tensor([M, M + kv_cache]), Tensor([M + kv_cache, N]))
    mul2_fusion = MatmulFusion([mul2], data_type_dict['fp16'])
    mul_fusion = [copy.deepcopy(mul2_fusion) for _ in range(batch)]
    mul2 = HorizontalMatmulFusion(mul_fusion, data_type_dict['fp16'])
    SV_times = mul2.compile_and_simulate(system.device)
    SV_clock = SV_times * system.device.compute_module.clock_freq

    time_s = QK_times + S_times + SV_times
    clock = QK_clock + S_clock + SV_clock

    return time_s, area, clock

def process_config(args):
    """单独处理一个硬件配置的函数，用于多进程"""
    hardware_config, b, l, device_count, i, test_function = args
    hardware_config['config'] = config[i]
    hardware_config['core_count'] = core_count[i]
    hardware_config['sublane_count'] = sublane_count
    hardware_config['vector_width'] = vector_width[i]
    hardware_config['array_height'] = array_height[i]
    hardware_config['array_width'] = array_width[i]
    hardware_config['SRAM_KB'] = SRAM_KB[i]
    time_s, area, clock = test_function(hardware_config, b, l, device_count)
    print(f"Decode no V Config: {config[i]}, Batch Size: {b}, Lin: {l}, pass")
    return config[i], b, l, time_s, area, clock  # 返回更多信息

def save_to_xlsx(file_path, sheet_name, results, L, batch_size):
    cols = []
    cols.append(L)
    cols.append(batch_size)
    if "decode" in sheet_name:
        columns = pd.MultiIndex.from_product(cols, names=['Lout', 'batch_size'])
        index = config
    elif "prefill" in sheet_name:
        columns = pd.MultiIndex.from_product(cols, names=['Lin', 'batch_size'])
        index = config[:-1]
    else:
        Exception("sheet_name error")
    
    df = pd.DataFrame(index=index, columns=columns)
    for result in results:
        config_name, batch, Lout, time_s, area, clock = result
        df.loc[config_name, (Lout, batch)] = clock
        print(f"Decode Config: {config_name}, Batch Size: {batch}, Lin: {Lout}, Time: {time_s}, area: {area},Clock: {clock}")
    csv_file_path = file_path

    try:
        # 加载 Excel 文件
        book = load_workbook(csv_file_path)
        # 检查工作表是否存在
        if sheet_name in book.sheetnames:
            if len(book.sheetnames) == 1:
                book.create_sheet("temp_sheet")
            # 若存在，删除该工作表
            del book[sheet_name]
        # 保存修改后的 Excel 文件
        book.save(csv_file_path)

        with pd.ExcelWriter(csv_file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                # 将 DataFrame 写入指定工作表
            df.to_excel(writer, sheet_name=sheet_name)
    except FileNotFoundError:
        # 若文件不存在，直接将 DataFrame 写入新的 Excel 文件
        df.to_excel(csv_file_path, sheet_name=sheet_name)
        print(f"文件 {csv_file_path} 不存在，已创建并将 DataFrame 写入 {sheet_name} 工作表。")

    # 保存为 CSV 文件
    # df.to_csv(csv_file_path)
    read_df = pd.read_excel(csv_file_path, header=[0, 1], index_col=0, sheet_name=sheet_name)
    print("\n按原格式读取后的 DataFrame：")
    print(read_df)   
        


if __name__ == '__main__':
    hardware_config = {
        "config": "A100",
        'core_count': 128,
        "sublane_count": 4,
        "array_width": 16,
        "array_height": 16,
        'vector_width': 32,
        'SRAM_KB': 192,
    }
    
    batch_size = [1, 4, 16, 64, 256]
    Lin = [128, 1024, 2048]
    Lout = [128, 1024, 2048, 4096]
    device_count = 4

    time_s, area, clock = change_config_prefill_test(hardware_config, 1, 128, device_count)
    print("time_s:", time_s, "clock:",clock )
    file_path = "../ae/res1.xlsx"
    
    ## 构造任务列表
    tasks = []
    for l in Lin:
        for b in batch_size:
            for i in range(4):
                tasks.append((hardware_config.copy(), b, l, device_count, i, change_config_prefill_test))
    
    # 使用多进程池并行处理
    with Pool(processes=cpu_count()) as pool:
        results = pool.map(process_config, tasks)

    csv_file_path = "../ae/res1.xlsx"
    sheet_name = "prefill"
    save_to_xlsx(csv_file_path, sheet_name, results, Lin, batch_size)
    
    
    #  构造任务列表
    tasks = []
    for l in Lout:
        for b in batch_size:
            for i in range(5):
                tasks.append((hardware_config.copy(), b, l, device_count, i, change_config_decode_test))

    # 使用多进程池并行处理
    with Pool(processes=cpu_count()) as pool:
        results = pool.map(process_config, tasks)
        
    # # 定义 CSV 文件路径
    csv_file_path = "../ae/res1.xlsx"
    sheet_name = "decode"
    save_to_xlsx(csv_file_path, sheet_name, results, Lout, batch_size)
   
    #  构造任务列表
    tasks = []
    for l in Lout:
        for b in batch_size:
            for i in range(5):
                tasks.append((hardware_config.copy(), b, l, device_count, i, change_config_decode_no_V_test))
    
    
    # 使用多进程池并行处理
    with Pool(processes=cpu_count()) as pool:
        results = pool.map(process_config, tasks)
    
    # # 定义 CSV 文件路径
    csv_file_path = "../ae/res1.xlsx"
    sheet_name = "decode_no_V"
    save_to_xlsx(csv_file_path, sheet_name, results, Lout, batch_size)



    #
    # # 处理结果
    # for result in results:
    #     config_name, batch_size, Lout, time_s, area, clock = result
    #     print(
    #         f"Decode no V Config: {config_name}, Batch Size: {batch_size}, Lin: {Lout}, Time: {time_s}, area: {area},Clock: {clock}")

    