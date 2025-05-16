import sys
import os
import copy
import pandas as pd
# 获取当前工作目录
if __name__ == "__main__":
    current_dir = os.getcwd()
    print("当前目录:", current_dir)

    # 更改到上两级目录
    parent_dir = os.path.dirname(os.path.dirname(current_dir))
    os.chdir(parent_dir)
    sys.path.append(parent_dir)

    # 验证当前工作目录
    new_dir = os.getcwd()
    print("更改后的目录:", new_dir)
# sys.path.append("/root/paper/LLMfusion")
# os.chdir("/root/paper/LLMfusion")
import logging
from ae.fig1.change_size import change_hardware_params
from software_model.matmul_horizontal_fusion import HorizontalMatmulFusion
from software_model.mutmul_fusion import MatmulFusion
from software_model.matmul import Matmul, BatchedMatmul
from software_model.DataFrame import DataType, Tensor, data_type_dict
from hardware_model.system import System
from software_model.softmax import Softmax
from software_model.flash_attention_fusion import FlashAttentionFusion
from LLM_model.opt175b import opt175b_prefill, opt175b_decode
import json
import time
from loguru import logger
from util.mapping import Mapping
from multiprocessing import Pool, cpu_count
from openpyxl import load_workbook

config = ["LLMCompass_Latency", "LLMCompass_Throught", "GA100", "me_prefill"]
core_count = [64, 64, 128, 32]
sublane_count = [4, 4, 4,4]
vector_width = [32, 32, 32, 32]
array_height = [16, 32, 16, 64]
array_width = [16, 32, 16, 64]
SRAM_KB = [192, 768, 192, 1024]
global_buffer_MBS = [24, 48, 48, 48]

global_buffer_bandwidth_per_cycle_bytes = [2560, 5120, 5120, 5120]
memory_bandwidths = [5, 2.5, 5, 2.5]
total_capacity_GBS = [80, 512, 80, 80]
memory_protocols = ["HBM2e", "PCIe5", "HBM2e", "PCIe5"]
device_count = 4

def change_config_prefill_test(hardware_config, b, l):
    M = l
    prefill = opt175b_prefill(12288, 96, device_count, data_type=data_type_dict['fp16'])
    _ = prefill(Tensor([b, M, 12288]))
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(hardware_config, arch_specs)
    latency = prefill.compile_and_simulate(system)

    return latency, None

def change_config_decode_test(hardware_config, b, l):
    M = 1
    decode = opt175b_decode(12288, 96, device_count, data_type=data_type_dict['fp16'])
    _ = decode(Tensor([b, M, 12288]), l)
    with open('./configs/GA100.json', "r") as f:
        arch_specs = json.load(f)
    system, area = change_hardware_params(hardware_config, arch_specs)
    latency = decode.compile_and_simulate(system)

    return latency, None

def process_config(args):
    hardware_config, b, l, i, test_func = args
    hardware_config["config"] = config[i]
    hardware_config["core_count"] = core_count[i]
    hardware_config["sublane_count"] = sublane_count[i]
    hardware_config["vector_width"] = vector_width[i]
    hardware_config["array_width"] = array_width[i]
    hardware_config["array_height"] = array_height[i]
    hardware_config["SRAM_KB"] = SRAM_KB[i]
    hardware_config["global_buffer_MB"] = global_buffer_MBS[i]
    hardware_config["global_buffer_bandwidth_per_cycle_byte"] = global_buffer_bandwidth_per_cycle_bytes[i]
    hardware_config["memory_bandwidth"] = memory_bandwidths[i]
    hardware_config["total_capacity_GB"] = total_capacity_GBS[i]
    hardware_config["memory_protocol"] = memory_protocols[i]
    times, clock = test_func(hardware_config, b, l)
    print(f"Config: {config[i]}, batch: {b}, l: {l}, Latency: {times.get('total_latency', 0)}")
    return config[i], b, l, times

def save_to_xlsx(csv_file_path, sheet_name, results, L, B, C):
    index = list(results[0][-1].keys())
    cols = []
    cols.append(B)
    cols.append(L)
    cols.append(C)
    columns = pd.MultiIndex.from_product(cols, names=["Batch Size", "L", "Config"])

    df = pd.DataFrame(index=index, columns=columns)
    for result in results:
        config_name, b, l, values = result
        for k,v in values.items():
            df.loc[k, (b, l, config_name)] = v

    try:
        # 加载 Excel 文件
        book = load_workbook(csv_file_path)
        # 检查工作表是否存在
        if sheet_name in book.sheetnames:
            if len(book.sheetnames) == 1:
                book.create_sheet("temp_sheet")
            # 若存在，删除该工作表
            del book[sheet_name]
        # 保存修改后的 Excel 文件
        book.save(csv_file_path)
        with pd.ExcelWriter(csv_file_path, engine='openpyxl', mode='a', if_sheet_exists='new') as writer:
                # 将 DataFrame 写入指定工作表
            df.to_excel(writer, sheet_name=sheet_name)
    except FileNotFoundError:
        # 若文件不存在，直接将 DataFrame 写入新的 Excel 文件
        df.to_excel(csv_file_path, sheet_name=sheet_name)
        print(f"文件 {csv_file_path} 不存在，已创建并将 DataFrame 写入 {sheet_name} 工作表。")

    read_df = pd.read_excel(csv_file_path, header=[0, 1, 2], index_col=0, sheet_name=sheet_name)
    print("\n按原格式读取后的 DataFrame：")
    print(read_df)

           

if __name__ == "__main__":
    hardware_config = {
        "config": "A100",
        'core_count': 128,
        "sublane_count": 4,
        "array_width": 16,
        "array_height": 16,
        'vector_width': 32,
        'SRAM_KB': 192,
        'global_buffer_MB': 48,
        'global_buffer_bandwidth_per_cycle_byte': 5120,
        'memory_bandwidth': 5,
        'total_capacity_GB': 80,
        'memory_protocol': "HBM2",
    }
    batch_size = [1, 4, 8, 16]
    Lin = [256, 512, 1024, 1536, 2048]
    Lout = [256, 512,1024,2048, 4096]
    for i in range(len(Lout)):
        Lout[i] = Lout[i] -1 

    start_time = time.time()
    task =[]

    for i in range(len(config)):
        for b in batch_size:
            for l in Lin:
                task.append((copy.deepcopy(hardware_config), b, l, i, change_config_prefill_test))

    with Pool(processes=cpu_count()) as pool:
        results = pool.map(process_config, task)

    csv_file_path = "../ae/res5.xlsx"
    sheet_name = "prefill"
    save_to_xlsx(csv_file_path, sheet_name, results, Lin, batch_size, config)
    end_time = time.time()
    logger.debug(f"Prefill test time: {end_time - start_time} seconds")

    
    # tasks = []
    # for i in range(len(config)):
    #     for b in batch_size:
    #         for l in Lout:
    #             tasks.append((copy.deepcopy(hardware_config), b, l, i, change_config_decode_test))


    # # 使用多进程池并行处理
    # with Pool(processes=cpu_count()) as pool:
    #     results = pool.map(process_config, tasks)

    # # # 定义 CSV 文件路径
    # csv_file_path = "../ae/res5.xlsx"
    # sheet_name = "decode"
    # save_to_xlsx(csv_file_path, sheet_name, results, Lout, batch_size)
    # end_time = time.time()
    # logger.debug(f"Prefill test time: {end_time - start_time} seconds")

